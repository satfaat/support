import openpyxl


companies_old = []
companies_new = []

wb_old = openpyxl.load_workbook(filename = 'myb.xlsx')
wb_new = openpyxl.load_workbook(filename = 'new.xlsx')

#print(wb_new.sheetnames)

sheet_old = wb_old['Лист1']
sheet_new = wb_new['TDSheet']
#print(sheet['B1'].value)
#print(sheet.cell(row=1, column=2).value)


for r in range(2, 43195): #43195
    #print(r, sheet.cell(row=r, column=2).value)
    companies_old.append(sheet_old.cell(row=r, column=2).value)

for r in range(3, 39329): #39329
    #print(r, sheet.cell(row=r, column=2).value)
    companies_new.append(sheet_new.cell(row=r, column=2).value)

#print(set(companies_old)-set(companies_new))
print(set(companies_new)-set(companies_old))


def get_xls_data():
    """arg = dict, pass value by key
    """

    path_xls = 'myb.xlsx'
    wb_old = openpyxl.load_workbook(filename=path_xls)

    for r in range(2, 43195): #43195
        #print(r, sheet.cell(row=r, column=2).value)
        companies_old.append(sheet_old.cell(row=r, column=2).value)
